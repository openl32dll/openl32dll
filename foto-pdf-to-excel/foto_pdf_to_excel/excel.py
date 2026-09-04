"""Tablo verilerini biçimlendirilmiş bir Excel (.xlsx) dosyasına yazar."""

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

Tablo = List[List[str]]


def _sayfa_adi_temizle(ad: str, kullanilan_adlar: set) -> str:
    """Excel sayfa adlarındaki geçersiz karakterleri temizler ve benzersizliği sağlar."""
    gecersiz = set('[]:*?/\\')
    temiz = "".join(c for c in ad if c not in gecersiz).strip()[:31] or "Sayfa"

    orijinal = temiz
    sayac = 2
    while temiz in kullanilan_adlar:
        ek = f" ({sayac})"
        temiz = orijinal[: 31 - len(ek)] + ek
        sayac += 1

    kullanilan_adlar.add(temiz)
    return temiz


def tablolari_excele_yaz(sayfalar: Dict[str, Tablo], cikti_yolu: str, ilk_satir_baslik: bool = True) -> None:
    """Her tabloyu ayrı bir Excel sayfası (worksheet) olarak dosyaya yazar.

    Parametreler
    ------------
    sayfalar:
        ``{sayfa_adi: [[hucre, hucre, ...], ...]}`` biçiminde sözlük.
    cikti_yolu:
        Yazılacak .xlsx dosyasının yolu.
    ilk_satir_baslik:
        ``True`` ise her sayfanın ilk satırı kalın/renkli başlık olarak
        biçimlendirilir ve dondurulur (freeze panes).
    """
    kitap = Workbook()
    kitap.remove(kitap.active)
    kullanilan_adlar: set = set()

    for sayfa_adi, tablo in sayfalar.items():
        ws = kitap.create_sheet(_sayfa_adi_temizle(sayfa_adi, kullanilan_adlar))

        if not tablo:
            continue

        for satir_no, satir in enumerate(tablo, start=1):
            for sutun_no, deger in enumerate(satir, start=1):
                ws.cell(row=satir_no, column=sutun_no, value=deger)

        if ilk_satir_baslik:
            for hucre in ws[1]:
                hucre.font = Font(bold=True, color="FFFFFF")
                hucre.fill = PatternFill("solid", fgColor="4472C4")
            ws.freeze_panes = "A2"

        sutun_sayisi = max(len(satir) for satir in tablo)
        for sutun_no in range(1, sutun_sayisi + 1):
            en_uzun = max(
                (len(str(satir[sutun_no - 1])) for satir in tablo if len(satir) >= sutun_no),
                default=10,
            )
            ws.column_dimensions[get_column_letter(sutun_no)].width = min(max(en_uzun + 2, 10), 60)

    if not kitap.sheetnames:
        kitap.create_sheet("Sayfa1")

    kitap.save(cikti_yolu)
